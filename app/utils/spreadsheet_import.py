"""
Spreadsheet import utility.
Reads a printer inventory xlsx file and upserts records into printer_import_data
and creates Location rows for each processed sheet.

Supported sheets: PMBS, Warehouse, Think Tank, HRPayroll, Wadley
Skipped: Medcomm (different schema), any *Diagram sheets, PMBS HEATERS
"""
from __future__ import annotations

import io
import logging
import re
from typing import Optional

logger = logging.getLogger(__name__)

# Sheets to process (case-insensitive match)
LOCATION_SHEETS = {"pmbs", "warehouse", "think tank", "hrpayroll", "wadley"}
# Sheets to always skip
SKIP_SHEETS = {"medcomm", "pmbs heaters"}

# Column name → field mapping (lowercase, stripped)
COLUMN_MAP = {
    "printer i.p.": "ip_address",
    "printer ip":   "ip_address",
    "ip address":   "ip_address",
    "ip":           "ip_address",
    "person":       "assigned_person",
    "name":         "assigned_person",
    "computer":     "assigned_computer",
    "pc":           "assigned_computer",
    "ext":          "phone_ext",
    "extension":    "phone_ext",
    "sql number":   "sql_number",
    "sql #":        "sql_number",
    "sql#":         "sql_number",
    "asset":        "sql_number",
    "printer username": "printer_web_username",
    "printer user":     "printer_web_username",
    "username":         "printer_web_username",
    "printer password": "printer_web_password",
    "printer pass":     "printer_web_password",
    "password":         "printer_web_password",
}

_IP_RE = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")


def _clean_ip(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ("", "n/a", "na", "none", "-"):
        return None
    return s if _IP_RE.match(s) else None


def _str(val) -> Optional[str]:
    if val is None:
        return None
    # openpyxl reads numeric cells as float; convert whole numbers to int first
    # so that str(5.0) becomes "5" instead of "5.0"
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    s = str(val).strip()
    return s if s and s.lower() not in ("n/a", "na", "none", "-") else None


def import_printer_spreadsheet(file_bytes: bytes) -> dict:
    """
    Parse xlsx bytes, upsert into printer_import_data and create Location rows.
    Returns {"imported": N, "skipped": N, "locations": [...], "errors": [...]}
    """
    import openpyxl
    from app.core.database import db
    from app.models.location import Location
    from app.models.printer_import import PrinterImportData
    from sqlalchemy.dialects.mysql import insert as mysql_insert

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return {"imported": 0, "skipped": 0, "locations": [], "errors": [f"Could not open file: {e}"]}

    imported = 0
    skipped = 0
    errors = []
    processed_locations = []

    # ip → dict of accumulated values (for shared-printer person merging)
    staging: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        normalized = sheet_name.strip().lower()

        # Skip diagram sheets and explicitly excluded sheets
        if "diagram" in normalized or normalized in SKIP_SHEETS:
            logger.debug("Skipping sheet: %s", sheet_name)
            continue

        # Only process known location sheets
        if normalized not in LOCATION_SHEETS:
            logger.debug("Skipping unknown sheet: %s", sheet_name)
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Build column index map from header row
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        col_idx: dict[str, int] = {}
        for i, h in enumerate(header):
            field = COLUMN_MAP.get(h)
            if field and field not in col_idx:
                col_idx[field] = i

        if "ip_address" not in col_idx:
            errors.append(f"Sheet '{sheet_name}': no IP address column found — skipped.")
            continue

        location_name = sheet_name.strip()
        processed_locations.append(location_name)

        for row in rows[1:]:
            if all(c is None for c in row):
                continue

            def get(field: str):
                idx = col_idx.get(field)
                return row[idx] if idx is not None else None

            ip = _clean_ip(get("ip_address"))
            if not ip:
                skipped += 1
                continue

            person = _str(get("assigned_person"))

            if ip in staging:
                # Shared printer — merge person names
                existing = staging[ip]
                if person and person not in (existing.get("assigned_person") or ""):
                    prev = existing.get("assigned_person") or ""
                    existing["assigned_person"] = f"{prev}, {person}".lstrip(", ")
            else:
                staging[ip] = {
                    "ip_address":           ip,
                    "location_name":        location_name,
                    "assigned_person":      person,
                    "sql_number":           _str(get("sql_number")),
                    "assigned_computer":    _str(get("assigned_computer")),
                    "phone_ext":            _str(get("phone_ext")),
                    "printer_web_username": _str(get("printer_web_username")),
                    "printer_web_password": _str(get("printer_web_password")),
                }

    wb.close()

    # Upsert Location rows
    for loc_name in processed_locations:
        existing_loc = db.session.query(Location).filter_by(name=loc_name).first()
        if not existing_loc:
            db.session.add(Location(name=loc_name))
    try:
        db.session.flush()
    except Exception as e:
        errors.append(f"Location upsert error: {e}")

    # Upsert PrinterImportData rows
    for ip, rec in staging.items():
        try:
            existing = db.session.get(PrinterImportData, ip)
            if existing:
                for k, v in rec.items():
                    if k != "ip_address":
                        setattr(existing, k, v)
            else:
                db.session.add(PrinterImportData(**rec))
            imported += 1
        except Exception as e:
            errors.append(f"Row {ip}: {e}")
            skipped += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return {"imported": 0, "skipped": skipped, "locations": [], "errors": [f"DB commit failed: {e}"]}

    return {
        "imported": imported,
        "skipped": skipped,
        "locations": processed_locations,
        "errors": errors,
    }
